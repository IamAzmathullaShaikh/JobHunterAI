import io
import pandas as pd
from typing import List, Optional
from sqlalchemy import select
from sqlalchemy.orm import selectinload
from sqlalchemy.ext.asyncio import AsyncSession
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from database.models import JobListing, AIAnalysis, JobApplication, ApplicationStatus
from utils.logger import logger

class ExportService:
    def __init__(self, session: AsyncSession):
        self.session = session

    async def fetch_jobs_dataframe(self, status_filter: Optional[str] = None) -> pd.DataFrame:
        """Queries database using SQLAlchemy 2.0 schema and builds a clean Pandas DataFrame."""
        stmt = select(JobListing).options(
            selectinload(JobListing.ai_analysis),
            selectinload(JobListing.application)
        )
        
        # Filter by ApplicationStatus enum if a specific status is chosen
        if status_filter and status_filter != "All":
            target_enum = None
            for status_item in ApplicationStatus:
                if status_item.value.lower() == status_filter.lower():
                    target_enum = status_item
                    break
            
            if target_enum:
                stmt = stmt.join(JobListing.application).where(JobApplication.status == target_enum)

        result = await self.session.execute(stmt)
        jobs: List[JobListing] = result.scalars().all()

        rows = []
        for j in jobs:
            # Safely resolve AI Analysis metrics
            match_score = j.ai_analysis.match_score if j.ai_analysis else 0.0
            fit_summary = j.ai_analysis.fit_summary if j.ai_analysis else "Pending Analysis"
            
            # Safely resolve Application Status Enum
            status_val = j.application.status.value if j.application else ApplicationStatus.IDENTIFIED.value
            
            rows.append({
                "Job ID": j.job_id_raw or "",
                "Status": status_val.upper(),
                "Match Score (%)": round(match_score, 1),
                "Job Title": j.title or "",
                "Company": j.company_name or "",
                "Location": j.location or "",
                "Source Portal": j.source or "",
                "Work Mode": j.work_place_type or "Onsite",
                "Employment Type": j.job_type or "Full-Time",
                "Application URL": j.url or "",
                "Captured Date": j.date_scraped.strftime("%Y-%m-%d %H:%M") if j.date_scraped else "",
                "AI Alignment Summary": fit_summary
            })

        df = pd.DataFrame(rows)
        
        if not df.empty:
            # Primary sort: AI Match Score descending; Secondary sort: Capture date
            df = df.sort_values(by=["Match Score (%)", "Captured Date"], ascending=[False, False]).reset_index(drop=True)
            
        return df

    async def generate_styled_excel(self, status_filter: Optional[str] = None) -> bytes:
        """Generates a polished, minimal Excel spreadsheet bytes stream with openpyxl styling."""
        df = await self.fetch_jobs_dataframe(status_filter=status_filter)
        
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            sheet_name = "Job Tracker"
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            
            workbook = writer.book
            worksheet = writer.sheets[sheet_name]
            
            # Palette Definition
            font_family = "Segoe UI"
            header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid") # Dark Charcoal
            header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
            
            body_font = Font(name=font_family, size=10, color="1F2937")
            link_font = Font(name=font_family, size=10, color="2563EB", underline="single")
            
            # Status Badge Fills
            fill_identified = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid") # Light Gray
            fill_ai_ready = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")  # Soft Amber
            fill_applied = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")   # Soft Mint
            fill_interview = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid") # Soft Sky Blue
            fill_offer = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")     # Vibrant Mint
            fill_rejected = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")  # Soft Coral
            
            thin_border = Border(
                left=Side(style='thin', color='E5E7EB'),
                right=Side(style='thin', color='E5E7EB'),
                top=Side(style='thin', color='E5E7EB'),
                bottom=Side(style='thin', color='E5E7EB')
            )

            # Style Header Row
            for col_num in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=1, column=col_num)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            # Style Body Rows
            for row_num in range(2, len(df) + 2):
                status_val = str(worksheet.cell(row=row_num, column=2).value).upper()
                
                for col_num in range(1, len(df.columns) + 1):
                    cell = worksheet.cell(row=row_num, column=col_num)
                    cell.font = body_font
                    cell.border = thin_border
                    cell.alignment = Alignment(vertical="center")

                    # Highlight Status Badge Column (Column 2)
                    if col_num == 2:
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                        if "APPLIED" in status_val:
                            cell.fill = fill_applied
                        elif "INTERVIEW" in status_val:
                            cell.fill = fill_interview
                        elif "OFFER" in status_val:
                            cell.fill = fill_offer
                        elif "AI READY" in status_val:
                            cell.fill = fill_ai_ready
                        elif "REJECTED" in status_val:
                            cell.fill = fill_rejected
                        else:
                            cell.fill = fill_identified

                    # Center Aligned Metadata Columns
                    if col_num in [1, 3, 7, 8, 9, 11]:
                        cell.alignment = Alignment(horizontal="center", vertical="center")

                    # Clickable Application URL Hyperlinks (Column 10)
                    if col_num == 10 and cell.value and str(cell.value).startswith("http"):
                        cell.hyperlink = cell.value
                        cell.font = link_font

            # Enable Filters Across Column Headers
            worksheet.auto_filter.ref = worksheet.dimensions

            # Auto-fit Column Widths (min 12, max 50 chars)
            for col in worksheet.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    val_str = str(cell.value or '')
                    if len(val_str) > max_len:
                        max_len = len(val_str)
                worksheet.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 50)

        output.seek(0)
        logger.success("Generated polished Excel tracking workbook bytes matching DeclarativeBase schema.")
        return output.getvalue()